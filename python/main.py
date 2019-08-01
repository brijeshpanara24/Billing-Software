import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from tkinter import *
from tkinter import filedialog

#------------------------------- Fetch Customer Details ----------------------------------------------

wb = openpyxl.load_workbook('customer.xlsm')
sheet = wb.get_sheet_by_name('customer')

pos1 = []
Name = []
Adress1 = []
Adress2 = []
FileNickname = []
GSTNumber = []

for i in range(3,50):
	if sheet['A'+str(i)].value!=None and sheet['A'+str(i)].value!='':
		pos1.append(i-2)
		Name.append(sheet['A'+str(i)].value)
		Adress1.append(sheet['B'+str(i)].value)
		Adress2.append(sheet['C'+str(i)].value)
		GSTNumber.append(sheet['D'+str(i)].value)
		FileNickname.append(sheet['E'+str(i)].value)
		print([pos1[i-3],Name[i-3],Adress1[i-3],Adress2[i-3],GSTNumber[i-3],FileNickname[i-3]])

print('')
wb.close()

#---------------------------------- Fetch Item Details ---------------------------------------

wb = openpyxl.load_workbook('item.xlsm')
sheet = wb.get_sheet_by_name('item')

pos2 = []
Item = []

for i in range(3,50):
	if sheet['A'+str(i)].value!=None:
		pos2.append(i-2)
		Item.append(sheet['A'+str(i)].value)
		print([pos2[i-3],Item[i-3]])

print('')
wb.close()	

#------------------------------------- Autocomplete Class --------------------------------------------------

class AutocompleteEntry(Entry):
    def __init__(self, lista, *args, **kwargs):
        
        Entry.__init__(self, *args, **kwargs)
        self.lista = lista        
        self.var = self["textvariable"]
        if self.var == '':
            self.var = self["textvariable"] = StringVar()

        self.var.trace('w', self.changed)
        self.bind("<Right>", self.selection)
        self.bind("<Return>", self.selection)
        self.bind("<Up>", self.up)
        self.bind("<Down>", self.down)
        
        self.lb_up = False

    def changed(self, name, index, mode):  

        if self.var.get() == '':
            self.lb.destroy()
            self.lb_up = False
        else:
            words = self.comparison()
            if words:            
                if not self.lb_up:
                    self.lb = Listbox(width = 42)
                    self.lb.bind("<Double-Button-1>", self.selection)
                    self.lb.bind("<Right>", self.selection)
                    self.lb.place(x=self.winfo_x(), y=self.winfo_y()+self.winfo_height())
                    self.lb_up = True
                
                self.lb.delete(0, END)
                for w in words:
                    self.lb.insert(END,w)
            else:
                if self.lb_up:
                    self.lb.destroy()
                    self.lb_up = False
        
    def selection(self, event):

        if self.lb_up:
            self.var.set(self.lb.get(ACTIVE))
            self.lb.destroy()
            self.lb_up = False
            self.icursor(END)

    def up(self, event):

        if self.lb_up:
            if self.lb.curselection() == ():
                index = '0'
            else:
                index = self.lb.curselection()[0]
            if index != '0':                
                self.lb.selection_clear(first=index)
                index = str(int(index)-1)                
                self.lb.selection_set(first=index)
                self.lb.activate(index) 
    

    def down(self, event):

        if self.lb_up:
            if self.lb.curselection() == ():
                index = '-1'
            else:
                index = self.lb.curselection()[0]
            if index != END:                        
                self.lb.selection_clear(first=index)
                index = str(int(index)+1)        
                self.lb.selection_set(first=index)
                self.lb.activate(index) 
            
    def comparison(self):
    	words = []
    	pattern = self.var.get().lower()
    	for str in self.lista:
    		if pattern in str.lower():
    			words.append(str)
    	return words


#------------------------------------ Designing Excel File ---------------------------------------------

medium_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))

top_medium_border = Border( top=Side(style='medium'))
bottom_medium_border = Border( bottom=Side(style='medium'))

left_medium_border = Border( left=Side(style='medium'))
right_medium_border = Border( right=Side(style='medium'))

top_left_medium_border = Border(left=Side(style='medium'),top=Side(style='medium'))
top_right_medium_border = Border(right=Side(style='medium'),top=Side(style='medium'))
bottom_left_medium_border = Border(left=Side(style='medium'),bottom=Side(style='medium'))
bottom_right_medium_border = Border(right=Side(style='medium'),bottom=Side(style='medium'))


def design_excel_file():
	#design
	row = [8,0,0,0,0,0,0,8,0,0,0,0,8,4,8,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,4,4,4,4,4,8,8,0,0,0,0,0,8]
	rowOffset = [0,0,0,0,0,0,0,0,0,0,0,0,0,4,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,4,4,4,4,4,0,0,0,0,0,0,0,0]
	colmn = [48,28,28,33,29,0,29,48]
	colmnOffset = [0,12,12,7,12,0,12,0]
	rowColmnTopLeft = [[1,8,8,13,13,13,13,13,13,14,14,15,15,15,15,15,15,36,36,37,37,38,38,39,39,40,40,41,41,41,42],
					   [1,1,4,1,2,3,4,5,7,5,7,1,2,3,4,5,7,5,7,5,7,5,7,5,7,5,7,1,5,7,1]]
	rowColmnTopRight = [[1,8,13,14,15,36,37,38,39,40,41,42],
					    [8,8,8,8,8,8,8,8,8,8,8,8]]

	for i in range(1,len(row)+1):
		for j in range(1,row[i-1]+1):
			if i!=len(row):
				sheet.cell(row=i, column=rowOffset[i-1]+j).border = top_medium_border
			else:
				sheet.cell(row=i, column=rowOffset[i-1]+j).border = bottom_medium_border	

	for i in range(1,len(colmn)+1):
		for j in range(1,colmn[i-1]+1):
			if i!=len(colmn):
				sheet.cell(row=colmnOffset[i-1]+j, column=i).border = left_medium_border
			else:
				sheet.cell(row=colmnOffset[i-1]+j, column=i).border = right_medium_border		
	
	for i in range(0,len(rowColmnTopLeft[0])):
		sheet.cell(row=rowColmnTopLeft[0][i],column=rowColmnTopLeft[1][i]).border = top_left_medium_border

	for i in range(0,len(rowColmnTopRight[0])):
		sheet.cell(row=rowColmnTopRight[0][i],column=rowColmnTopRight[1][i]).border = top_right_medium_border

	sheet.cell(row=48,column=8).border = bottom_right_medium_border
	sheet.cell(row=48,column=1).border = bottom_left_medium_border

#-------------------------------------- Save Excel File Data --------------------------------------------------------------

def fill_excel_file():
	global toplevel
	on_closing_toplevel()

	filenickname = None
	
	#customerName
	if (customer.get() not in Name) or (date1.get()==None or date1.get()=='') or (billNumber.get()==None or billNumber.get()==''):
		toplevel = Tk()
		toplevel.title("Error")
		toplevel.resizable(FALSE,FALSE)
		toplevel.protocol("WM_DELETE_WINDOW", on_closing_toplevel)
		templabel = Label(toplevel, text="\n\n\n\nInvalid Data\n\nPlease verify all details\n\n")
		templabel.pack()
		tempbtn = Button(toplevel, text='Ok', command=on_closing_toplevel)
		tempbtn.pack()
		w = 250
		h = 250
		x = master.winfo_x() + int((master.winfo_width()/2)) - int(250/2)
		y = master.winfo_y() + int((master.winfo_height()/2)) - int(250/2)
		toplevel.geometry('%dx%d+%d+%d' % (w, h, x, y))
		return filenickname

	response = Name.index(customer.get())

	sheet['B8'] = Name[response]
	sheet['B9'] = Adress1[response]
	sheet['B10'] = Adress2[response]
	sheet['B12'] = GSTNumber[response]
	filenickname = FileNickname[response]
	if filenickname==None:
		filenickname=''
		
	#date
	sheet['G9'] = date1.get()
	sheet['G10'] = date2.get()
	sheet['G11'] = date3.get()

	#billNumber
	if billNumber.get()!=None and billNumber.get()!='':
		sheet['E9'] = billNumber.get()
	if challanNumber.get()!=None and challanNumber.get()!='':
		sheet['E10'] = challanNumber.get()
	if orderNumber.get()!=None and orderNumber.get()!='':
		sheet['E11'] = orderNumber.get()

	#itemName
	for i in range(0,len(colmn1)):
		if colmn1[i].get()!=None and colmn1[i].get()!='':
			pos = i+15
			sheet['A'+str(pos)] = int(i+1)
			sheet['B'+str(pos)] = colmn1[i].get()
			if(colmn2[i].get()!=None and colmn2[i].get()!=''):
				sheet['C'+str(pos)] = colmn2[i].get()
			if(colmn3[i].get()!=None and colmn3[i].get()!=''):
				sheet['D'+str(pos)] = colmn3[i].get()
				currentCell = sheet['D'+str(pos)]
				currentCell.alignment = Alignment(horizontal='right')
			if(colmn4[i].get()!=None and colmn4[i].get()!=''):
				sheet['E'+str(pos)] = colmn4[i].get()
				currentCell = sheet['E'+str(pos)]
				currentCell.alignment = Alignment(horizontal='right')

	#GST
	if(GST.get()==1):
		sheet['G38']=0
		currentCell = sheet['G38']
		currentCell.alignment = Alignment(horizontal='right')
		sheet['G39']=0
		currentCell = sheet['G39']
		currentCell.alignment = Alignment(horizontal='right')
	else:
		sheet['G40']=0
		currentCell = sheet['G40']
		currentCell.alignment = Alignment(horizontal='right')

	design_excel_file()	 
	return filenickname

#--------------------------------------------------------------------------------------------------------

def save_excel_file():
	global wb,sheet

	on_closing_toplevel()

	dir = os.getcwd()	
	filenickname = fill_excel_file()

	if filenickname==None:
		return filenickname
	elif filenickname=='':
		filenickname = 'NewBill'
	
	if billNumber.get()!=None and billNumber.get()!='':
		filenickname = "BILL "+ billNumber.get() + " " +filenickname
	else:
		filenickname = "BILL "+ str(0) + " " +filenickname
	
	file = filedialog.asksaveasfile(mode='w',initialfile=filenickname,defaultextension=".xlsm",parent=master,initialdir = dir+"\BillBackup", title='Please select a directory')
	if file!=None:
		wb.save(file.name)
		wb.close()
		wb = load_workbook(filename='bill.xlsm', read_only=False, keep_vba=True)
		sheet = wb.get_sheet_by_name('bill')
		return file.name
	else:
		return None


def print_excel_file():
	on_closing_toplevel()

	filenickname = fill_excel_file()
	if filenickname!=None:
		file = os.getcwd()+'\PrintFile.xlsm'
		wb.save(file)
		os.startfile(file, "print")


def save_and_print_excel_file():
	on_closing_toplevel()

	file = save_excel_file()
	if file!=None:
		os.startfile(file, "print")	

def on_closing_toplevel():
	global toplevel
	if (toplevel is not None):
		toplevel.destroy()
		toplevel = None

def on_closing_master():
	on_closing_toplevel()
	wb.close()
	master.quit()

#------------------------------------ Opening Tkinter Window ---------------------------------------------

wb = load_workbook(filename='bill.xlsm', read_only=False, keep_vba=True)
sheet = wb.get_sheet_by_name('bill')
toplevel = None

master = Tk()
master.title('BILL')
master.resizable(FALSE,FALSE)
#master.geometry("1000x700") #Width x Height

rows = 0
while rows < 50:
    master.rowconfigure(rows, weight=1)
    master.columnconfigure(rows,weight=1)
    rows += 1

Label(master, text="Bill No. = ").grid(row=0,column=0,sticky=E,pady=20)
billNumber = Entry(master)
billNumber.grid(row=0, column=1,sticky=W,pady=20)

Label(master, text="Challan No. = ").grid(row=1,column=0,sticky=E,pady=(0,20))
challanNumber = Entry(master)
challanNumber.grid(row=1, column=1,sticky=W,pady=(0,20))

Label(master, text="Order No. = ").grid(row=2,column=0,sticky=E,pady=(0,20))
orderNumber = Entry(master)
orderNumber.grid(row=2, column=1,sticky=W,pady=(0,20))

Label(master, text="Date 1 = ").grid(row=0,column=3,sticky=E,pady=20)
date1 = Entry(master)
date1.grid(row=0, column=4,sticky=W,pady=20)

Label(master, text="Date 2 = ").grid(row=1,column=3, sticky=E,pady=(0,20))
date2 = Entry(master)
date2.grid(row=1, column=4, sticky=W,pady=(0,20))

Label(master, text="Date 3 = ").grid(row=2,column=3,sticky=E,pady=(0,20))
date3 = Entry(master)
date3.grid(row=2, column=4,sticky=W,pady=(0,20))

Label(master, text="Customer = ").grid(row=3,column=0,sticky=E,pady=(0,20))
customer = AutocompleteEntry(Name,master)
customer.grid(row=3,column=1,columnspan=2,sticky=W+E,pady=(0,20))

Label(master, text="Item").grid(row=4,column=1,sticky=E)
Label(master, text="Name").grid(row=4,column=2,sticky=W)
Label(master, text="Code").grid(row=4,column=3)
Label(master, text="Quantity").grid(row=4,column=4)
Label(master, text="Rate").grid(row=4,column=5)

#just for column width format
Entry(master).grid(row=14, column=2,sticky=E)
		
colmn1 = []
colmn2 = []
colmn3 = []
colmn4 = []
for i in range(1,16):
	Label(master, text="Item "+str(i)+" = ").grid(row=i+4,column=0,padx=(50,0),sticky=E)
	temp1=AutocompleteEntry(Item,master)
	temp1.grid(row=i+4,column=2,sticky=W+E)
	temp1.grid(row=i+4,column=1,columnspan=2,sticky=W+E)
	temp2 = Entry(master)
	temp2.grid(row=i+4, column=3,sticky=E)
	temp3 = Entry(master)
	temp3.grid(row=i+4, column=4,sticky=E)
	temp4 = Entry(master)
	temp4.grid(row=i+4, column=5,padx=(0,50),sticky=E)
	colmn1.append(temp1)
	colmn2.append(temp2)
	colmn3.append(temp3)
	colmn4.append(temp4)

GST = IntVar()
c1 = Checkbutton(master,text = "IGST",variable = GST,onvalue = 1,offvalue = 0)
c1.grid(row=21, column=1 ,pady=20)
c2 = Checkbutton(master,text = "SGST+CGST",variable = GST,onvalue = 0,offvalue = 1)
c2.grid(row=21, column=5 ,pady=20)

# btn = Button(master, text='Exit', command=master.quit)
# btn.grid(row=17, column=1,pady=(0,20))

btn1 = Button(master, text='Save', command=save_excel_file)
btn1.grid(row=22, column=1,pady=(0,20))

btn2 = Button(master, text='Print', command=print_excel_file)
btn2.grid(row=22, column=3,pady=(0,20))

btn3 = Button(master, text='Save And Print',command=save_and_print_excel_file)
btn3.grid(row=22, column=5,pady=(0,20))

master.protocol("WM_DELETE_WINDOW", on_closing_master)
mainloop( )

